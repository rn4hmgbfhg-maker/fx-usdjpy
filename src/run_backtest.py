# -*- coding: utf-8 -*-
"""バックテスト実行スクリプト

全戦略を全期間+期間分割(イン/アウトオブサンプル)で検証し、
results/バックテスト結果.xlsx にサマリ表とエクイティカーブ(Excelグラフ)を出力する。
使い方:  python3 src/run_backtest.py
"""
import os
import pandas as pd

from backtest import Backtester
from strategies import STRATEGIES

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSV_PATH = os.path.join(BASE_DIR, "data", "usdjpy_daily.csv")
OUT_XLSX = os.path.join(BASE_DIR, "results", "バックテスト結果.xlsx")

PERIODS = {
    "全期間": (None, None),
    "前半(イン)": (None, "2018-12-31"),
    "後半(アウト)": ("2019-01-01", None),
}


def slice_df(df, start, end):
    d = pd.to_datetime(df["Date"])
    mask = pd.Series(True, index=df.index)
    if start:
        mask &= d >= start
    if end:
        mask &= d <= end
    return df[mask].reset_index(drop=True)


def main():
    df = pd.read_csv(CSV_PATH)
    rows = []
    curves = {}   # 全期間のエクイティカーブ(グラフ用)
    trades_all = {}

    for pname, (start, end) in PERIODS.items():
        sub = slice_df(df, start, end)
        for sname, func in STRATEGIES.items():
            bt = Backtester(sub)
            res = bt.run(func(sub))
            m = res["metrics"]
            rows.append({"期間": pname, "戦略": sname, **m,
                         "キルスイッチ": "発動" if res["halted"] else "－"})
            if pname == "全期間":
                curves[sname] = res["equity_curve"]
                trades_all[sname] = res["trades"]

    summary = pd.DataFrame(rows)
    print(summary.to_string(index=False))

    write_report(summary, curves, trades_all)
    print(f"\nレポート出力: {OUT_XLSX}")


def write_report(summary: pd.DataFrame, curves: dict, trades_all: dict):
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ---- シート1: サマリ
    ws = wb.active
    ws.title = "サマリ"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Border(*[Side(style="thin", color="BFBFBF")] * 4)

    ws.append(list(summary.columns))
    for c in ws[1]:
        c.fill, c.font, c.border = header_fill, header_font, thin
        c.alignment = Alignment(horizontal="center")
    pct_cols = {"総リターン", "年率リターン(CAGR)", "最大ドローダウン", "勝率"}
    for _, r in summary.iterrows():
        ws.append(list(r))
    for j, col in enumerate(summary.columns, start=1):
        for i in range(2, ws.max_row + 1):
            cell = ws.cell(row=i, column=j)
            cell.border = thin
            if col in pct_cols:
                cell.number_format = "0.0%"
            elif col == "最終資産(円)":
                cell.number_format = "#,##0"
            elif col in ("シャープレシオ", "プロフィットファクター"):
                cell.number_format = "0.00"
        width = max(len(str(col)) * 2 + 2, 10)
        ws.column_dimensions[get_column_letter(j)].width = width
    ws.freeze_panes = "A2"

    # ---- シート2: エクイティカーブ(全期間・戦略比較)
    ws2 = wb.create_sheet("エクイティカーブ")
    names = list(curves.keys())
    base = curves[names[0]]
    ws2.append(["Date"] + names)
    merged = base[["Date"]].copy()
    for n in names:
        merged[n] = curves[n]["Equity"].values
    for _, r in merged.iterrows():
        ws2.append([str(r["Date"])] + [round(r[n]) for n in names])
    chart = LineChart()
    chart.title = "エクイティカーブ（初期資金100万円・全期間）"
    chart.y_axis.title = "口座資産(円)"
    chart.x_axis.title = "日付"
    chart.height, chart.width = 12, 30
    data = Reference(ws2, min_col=2, max_col=1 + len(names),
                     min_row=1, max_row=ws2.max_row)
    cats = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws2.add_chart(chart, "F2")

    # ---- シート3〜: 戦略別トレード履歴
    for n, tr in trades_all.items():
        safe = n.replace("/", "・").replace("\\", "・").replace(":", "").replace("?", "").replace("*", "").replace("[", "(").replace("]", ")")
        ws3 = wb.create_sheet(f"取引履歴_{safe[:20]}")
        cols = ["entry_date", "exit_date", "dir", "entry", "exit", "pnl", "reason"]
        jp = ["建玉日", "決済日", "売買", "建値", "決済値", "損益(円)", "決済理由"]
        ws3.append(jp)
        for c in ws3[1]:
            c.fill, c.font = header_fill, header_font
        if len(tr):
            for _, r in tr[cols].iterrows():
                ws3.append([str(r.entry_date), str(r.exit_date),
                            "買" if r.dir == "L" else "売",
                            round(r.entry, 3), round(r.exit, 3),
                            round(r.pnl), r.reason])
        for j in range(1, 8):
            ws3.column_dimensions[get_column_letter(j)].width = 13

    os.makedirs(os.path.dirname(OUT_XLSX), exist_ok=True)
    wb.save(OUT_XLSX)


if __name__ == "__main__":
    main()
