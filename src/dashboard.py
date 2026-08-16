# -*- coding: utf-8 -*-
"""Google Drive連動 運用ダッシュボード生成（2システム構成）

毎朝の最終確定シグナル後(および研究でパラメータが進化した後)に自動実行され、
Driveの専用フォルダへ「FXドル円_運用ダッシュボード.xlsx」を書き出す。

シート構成:
  1. ダッシュボード   … 両システムの本日状況・KPI・注文指示・直近履歴・研究状況
  2. パフォーマンス   … 資産推移グラフ(実運用が浅いうちは15年検証の合成カーブ)・月次表
  3. 日次ログ / 4. 取引履歴 / 5. 研究ログ … 明細
"""
import json
import os
import sys
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from backtest import Backtester  # noqa: E402
from strategies import donchian  # noqa: E402
import perf  # noqa: E402
import swap  # noqa: E402

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
R = lambda *p: os.path.join(BASE_DIR, *p)  # noqa: E731

# ---- パレット(落ち着いた金融ダッシュボード調)
NAVY = "1F4E79"; DEEP = "142F4C"; GOLD = "C9A227"; CARD = "F4F7FB"
BUY = "C0392B"; SELL = "2E5FA3"; FLAT = "7F8C8D"
PROFIT = "1E7E34"; LOSS = "C0392B"; GRAYTX = "666666"; LINE_ = "D9E2EC"

F_TITLE = Font(name="Yu Gothic", size=18, bold=True, color="FFFFFF")
F_SUB = Font(name="Yu Gothic", size=9, color="DDE6F0")
F_SECT = Font(name="Yu Gothic", size=12, bold=True, color=NAVY)
F_LABEL = Font(name="Yu Gothic", size=9, color=GRAYTX)
F_SMALL = Font(name="Yu Gothic", size=9, color=GRAYTX)
F_HEAD = Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF")
F_BODY = Font(name="Yu Gothic", size=10, color="333333")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Border(*[Side(style="thin", color=LINE_)] * 4)


def _read_csv(path, cols):
    if os.path.exists(path):
        try:
            df = pd.read_csv(path)
            if len(df):
                return df
        except Exception:
            pass
    return pd.DataFrame(columns=cols)


def _card(ws, row, col, label, value, color=DEEP):
    c1, c2 = get_column_letter(col), get_column_letter(col + 1)
    ws.merge_cells(f"{c1}{row}:{c2}{row}")
    ws.merge_cells(f"{c1}{row+1}:{c2}{row+2}")
    lab = ws[f"{c1}{row}"]; lab.value = label
    lab.font = F_LABEL; lab.alignment = CENTER
    val = ws[f"{c1}{row+1}"]; val.value = value
    val.font = Font(name="Yu Gothic", size=14, bold=True, color=color)
    val.alignment = CENTER
    for rr in range(row, row + 3):
        for cc in (col, col + 1):
            cell = ws.cell(row=rr, column=cc)
            cell.fill = PatternFill("solid", fgColor=CARD)
            cell.border = THIN


def _sect(ws, row, text):
    ws.cell(row=row, column=2, value=text).font = F_SECT


def build():
    cfg = json.load(open(R("config.json"), encoding="utf-8"))
    systems = {k: v for k, v in cfg.get("システム", {}).items()
               if v.get("有効", True)}
    state = json.load(open(R("state.json"), encoding="utf-8")) \
        if os.path.exists(R("state.json")) else {}
    if "position" in state:      # 旧形式
        state = {"長期トレンド": state}
    gd = cfg.get("Driveフォルダ")
    if not gd or not os.path.isdir(gd):
        print("Driveフォルダが見つからないためダッシュボード出力をスキップ")
        return None

    px = pd.read_csv(R("data", "usdjpy_daily.csv"))
    close = float(px["Close"].iloc[-1])
    daily = _read_csv(R("results", "daily_log.csv"),
                      ["日付", "システム", "終値", "シグナル", "アクション",
                       "ポジション", "ストップ", "数量", "実現損益累計",
                       "評価損益計", "モデル資産"])
    # 2026-08-14: trades.csv（日足2システム専用）だけを見ていたため、
    # デイトレ2系統の決済が成績に入っていなかった。Webボードと同じ
    # perf.load_all() で4システムを合算する。
    trades = perf.load_all()
    research = _read_csv(R("results", "research_log.csv"),
                         ["日付", "システム", "エントリー", "イグジット", "ATR係数",
                          "PF全", "PF前半", "PF後半", "シャープ", "勝率", "最大DD",
                          "取引数", "スコア", "判定"])

    # 各システムの現在シグナルと15年検証(理論値)
    sysinfo = {}
    ref_curves = {}
    for name, s in systems.items():
        n_e, n_x = int(s["エントリー期間"]), int(s["イグジット期間"])
        atr_k = float(s["ATRストップ係数"])
        sig = int(donchian(px, n=n_e, exit_n=n_x).iloc[-1])
        bt = Backtester(px, initial_capital=float(cfg["口座資金_円"]),
                        risk_per_trade=float(s["リスク率"]),
                        max_leverage=float(s["レバ上限"]), atr_stop_mult=atr_k)
        ref = bt.run(donchian(px, n=n_e, exit_n=n_x))
        sysinfo[name] = {"params": f"ドンチャン{n_e}/{n_x}・ATR×{atr_k}",
                         "sig": sig, "m": ref["metrics"]}
        ref_curves[name] = ref["equity_curve"]

    # 本日の指示書(最終確定を優先)
    order_text = "（本日の確定指示書はまだ生成されていません）"
    odir = R("results", "orders")
    if os.path.isdir(odir):
        files = sorted(f for f in os.listdir(odir) if f.endswith(".txt"))
        finals = [f for f in files if "確定" in f] or files
        if finals:
            order_text = open(os.path.join(odir, finals[-1]),
                              encoding="utf-8").read().strip()

    wb = Workbook()

    # ============================================================ 1. ダッシュボード
    ws = wb.active
    ws.title = "ダッシュボード"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    for i in range(2, 12):
        ws.column_dimensions[get_column_letter(i)].width = 13.5

    ws.merge_cells("B1:K1")
    t = ws["B1"]; t.value = "FX ドル円 半自動売買 運用ダッシュボード（2システム）"
    t.font = F_TITLE; t.alignment = CENTER
    for c in range(2, 12):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=NAVY)
    ws.row_dimensions[1].height = 34

    ws.merge_cells("B2:K2")
    s2 = ws["B2"]
    sys_desc = "　/　".join(f"{n}: {i['params']}" for n, i in sysinfo.items())
    s2.value = (f"更新: {datetime.now():%Y-%m-%d %H:%M}　|　終値: {close:.3f}円"
                f"　|　{sys_desc}　|　運用先: MATRIX TRADER（手動発注）")
    s2.font = F_SUB; s2.alignment = CENTER
    for c in range(2, 12):
        ws.cell(row=2, column=c).fill = PatternFill("solid", fgColor=DEEP)
    ws.row_dimensions[2].height = 18

    # --- 本日の状況(システム別カード)
    row = 4
    for name, info in sysinfo.items():
        _sect(ws, row, f"◆ 本日の状況｜{name}（{info['params']}）")
        st = state.get(name, {})
        pos = int(st.get("position", 0))
        sig_txt = "買" if info["sig"] > 0 else "売" if info["sig"] < 0 else "無"
        sig_col = BUY if info["sig"] > 0 else SELL if info["sig"] < 0 else FLAT
        pos_txt = (f"買 {int(st.get('units', 0)):,}通貨" if pos > 0
                   else f"売 {int(st.get('units', 0)):,}通貨" if pos < 0 else "なし")
        stop_txt = f"{float(st['stop']):.3f}円" if st.get("stop") else "－"
        last_act = "－"
        drow = daily[daily["システム"] == name] if len(daily) else daily
        if len(drow):
            last_act = str(drow.iloc[-1]["アクション"])
        _card(ws, row + 1, 2, "シグナル", sig_txt, color=sig_col)
        _card(ws, row + 1, 4, "モデルポジション", pos_txt,
              color=BUY if pos > 0 else SELL if pos < 0 else FLAT)
        _card(ws, row + 1, 6, "直近アクション",
              last_act if len(drow) else "運用開始前", color=DEEP)
        _card(ws, row + 1, 8, "ストップ水準", stop_txt, color=GOLD)
        _card(ws, row + 1, 10, "建値",
              f"{float(st['entry']):.3f}円" if st.get("entry") else "－",
              color=DEEP)
        row += 5

    # --- 運用成績(合算)
    _sect(ws, row, "◆ 運用成績（実運用モデル・全4システム合算）")
    if len(trades):
        pnl = trades["損益円"].astype(float)
        win_rate = (pnl > 0).mean()
        pf = (pnl[pnl > 0].sum() / -pnl[pnl <= 0].sum()
              if (pnl <= 0).any() and pnl[pnl <= 0].sum() < 0 else float("nan"))
        wr_txt, pf_txt = f"{win_rate:.0%}", (f"{pf:.2f}" if pf == pf else "－")
        cum = pnl.sum()
        cum_txt = f"{cum:+,.0f}円"
        cum_col = PROFIT if cum >= 0 else LOSS
        n_tr = len(trades)
    else:
        wr_txt = pf_txt = "－"; cum_txt = "運用開始前"; cum_col = FLAT; n_tr = 0
    # モデル資産は「口座資金＋4システムの累計実現＋4システムの含み」を
    # その場で計算する。日足ログのモデル資産は日足2システムだけの値で、
    # かつ8時台の確定時点で凍結されるため、Webボードの同名KPIと食い違う。
    # 2026-08-14: 両ボードの定義をここで完全に一致させる。
    sw_cfg = swap.load(cfg)
    unreal_all = 0.0
    open_states = [(state.get(n) or {}) for n in state]
    for sf in ("state_intraday.json", "state_intraday15.json"):
        if os.path.exists(R(sf)):
            open_states.append(
                (json.load(open(R(sf), encoding="utf-8")) or {}).get("state", {}))
    for st0 in open_states:
        p0 = int(st0.get("position", 0) or 0)
        if not p0 or not st0.get("entry"):
            continue
        u0 = float(st0.get("units") or 0)
        unreal_all += (u0 * (close - float(st0["entry"])) * p0
                       + swap.accrued(p0, u0, st0.get("entry_date"),
                                      sw=sw_cfg)[0])
    equity = float(cfg["口座資金_円"]) + float(pnl.sum() if len(trades) else 0) \
        + unreal_all
    equity_txt = f"{equity:,.0f}円"
    days = daily["日付"].nunique() if len(daily) else 0
    _card(ws, row + 1, 2, "モデル資産", equity_txt, color=DEEP)
    _card(ws, row + 1, 4, "累計実現損益", cum_txt, color=cum_col)
    _card(ws, row + 1, 6, "勝率（実運用）", wr_txt, color=DEEP)
    _card(ws, row + 1, 8, "決済済取引", f"{n_tr}回", color=DEEP)
    _card(ws, row + 1, 10, "運用日数", f"{days}日", color=DEEP)
    row += 5

    # --- 検証値(システム別)
    for name, info in sysinfo.items():
        m = info["m"]
        _sect(ws, row, f"◆ 15年検証の実力｜{name}（{info['params']}）")
        _card(ws, row + 1, 2, "総リターン", f"{m['総リターン']:+.1%}", color=PROFIT)
        _card(ws, row + 1, 4, "勝率", f"{m['勝率']:.0%}", color=DEEP)
        _card(ws, row + 1, 6, "PF", f"{m['プロフィットファクター']:.2f}", color=DEEP)
        _card(ws, row + 1, 8, "最大DD", f"{m['最大ドローダウン']:.1%}", color=LOSS)
        _card(ws, row + 1, 10, "取引数(年)",
              f"{m['取引回数'] / 15.5:.0f}回", color=DEEP)
        row += 5

    # --- 本日の注文指示
    _sect(ws, row, "◆ 本日の注文指示（最終確定・指標イベント欄含む）")
    top, bottom = row + 1, row + 12
    ws.merge_cells(f"B{top}:K{bottom}")
    box = ws.cell(row=top, column=2, value=order_text)
    box.font = Font(name="Yu Gothic", size=10, color="222222")
    box.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for rr in range(top, bottom + 1):
        for cc in range(2, 12):
            cell = ws.cell(row=rr, column=cc)
            cell.fill = PatternFill("solid", fgColor="FFFBEF")
            cell.border = THIN
    row = bottom + 2

    # --- 直近の記録
    _sect(ws, row, "◆ 直近の記録（両システム）")
    heads = ["日付", "システム", "シグナル", "アクション", "ポジション",
             "ストップ", "実現損益累計", "評価損益計", "モデル資産"]
    hr = row + 1
    for j, h in enumerate(heads, start=2):
        c = ws.cell(row=hr, column=j, value=h)
        c.font = F_HEAD; c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = CENTER; c.border = THIN
    tail = daily.tail(12)
    if len(tail):
        for i, (_, r) in enumerate(tail.iterrows(), start=hr + 1):
            vals = [r["日付"], r["システム"], r["シグナル"], r["アクション"],
                    r["ポジション"], r["ストップ"], r["実現損益累計"],
                    r["評価損益計"], r["モデル資産"]]
            for j, v in enumerate(vals, start=2):
                c = ws.cell(row=i, column=j, value=v)
                c.font = F_BODY; c.border = THIN; c.alignment = CENTER
        row = hr + len(tail) + 2
    else:
        ws.merge_cells(f"B{hr+1}:K{hr+1}")
        c = ws.cell(row=hr + 1, column=2,
                    value="運用開始前（次回の【最終確定】配信から自動記録されます）")
        c.font = F_SMALL; c.alignment = CENTER; c.border = THIN
        row = hr + 3

    # --- 研究・進化
    _sect(ws, row, "◆ スキーム研究・進化の状況")
    if len(research):
        latest = research.iloc[-1]
        adopted = research[research["判定"] == "採用"]
        msg = (f"最終研究日: {latest['日付']}　|　パラメータ進化 累計{len(adopted)}回　|　"
               f"直近: {latest.get('システム', '')} {latest['判定']}"
               f"（候補 ドンチャン{latest['エントリー']}/{latest['イグジット']}・"
               f"ATR×{latest['ATR係数']}　PF{latest['PF全']}・勝率{latest['勝率']}）")
    else:
        msg = ("毎朝8時の配信後に15年データで両システムのパラメータ探索を自動実行し、"
               "頑健性基準を満たした場合のみ進化します（研究ログ シート参照）")
    ws.merge_cells(f"B{row+1}:K{row+2}")
    c = ws.cell(row=row + 1, column=2, value=msg)
    c.font = F_BODY; c.alignment = LEFT
    for rr in (row + 1, row + 2):
        for cc in range(2, 12):
            ws.cell(row=rr, column=cc).fill = PatternFill("solid", fgColor=CARD)
            ws.cell(row=rr, column=cc).border = THIN

    ws.merge_cells(f"B{row+4}:K{row+4}")
    d = ws.cell(row=row + 4, column=2)
    d.value = ("※発注はMATRIX TRADERで本人が手動で行う（JFXは自動売買禁止）。"
               "検証値は過去データによるもので将来の成果を保証しない。"
               "高重要度指標イベント日は新規建てを自動見送り。")
    d.font = F_SMALL

    # ============================================================ 2. パフォーマンス
    ws2 = wb.create_sheet("パフォーマンス")
    ws2.sheet_view.showGridLines = False
    eq_daily = (daily.drop_duplicates("日付", keep="last")[["日付", "モデル資産"]]
                if len(daily) else daily)
    use_real = len(eq_daily) >= 5
    if use_real:
        # 2026-08-14: ここは daily_log の「モデル資産」列（日足2システムのみ・
        # スワップ遡及補正前の凍結値）をそのまま使っていたため、同じブックの
        # ダッシュボードKPI（4システム合算）と資産が食い違っていた
        # （957,505円 と 955,593円）。Webボードと同じ perf.equity_curve に統一。
        pts = perf.equity_curve(daily, float(cfg["口座資金_円"]),
                                equity_now=equity, trades=trades)
        src = pd.DataFrame({"Date": [d for d, _ in pts],
                            "Equity": [v for _, v in pts]})
        title = "資産推移（実運用モデル・全4システム合算）"
    else:
        # 両システムの検証カーブ合成(それぞれ初期資金起点のため差分を合算)
        names = list(ref_curves.keys())
        base_c = ref_curves[names[0]][["Date", "Equity"]].copy()
        init = float(cfg["口座資金_円"])
        eq = base_c["Equity"].astype(float) - init
        for n in names[1:]:
            eq = eq + (ref_curves[n]["Equity"].astype(float).values - init)
        src = pd.DataFrame({"Date": base_c["Date"], "Equity": eq + init})
        title = "資産推移（15年検証・両システム合成の理論値）"
    ws2["A1"] = "Date"; ws2["B1"] = "Equity"
    for i, (_, r) in enumerate(src.iterrows(), start=2):
        ws2.cell(row=i, column=1, value=str(r["Date"]))
        ws2.cell(row=i, column=2, value=float(r["Equity"]))
    chart = LineChart()
    chart.title = title
    chart.height, chart.width = 13, 32
    chart.y_axis.title = "口座資産(円)"
    data = Reference(ws2, min_col=2, min_row=1, max_row=ws2.max_row)
    cats = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws2.add_chart(chart, "D2")
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 12

    if use_real and len(src) >= 2:
        ws2["D30"] = "月次サマリ（実運用・全4システム合算）"
        ws2["D30"].font = F_SECT
        dd = src.copy()
        dd["月"] = dd["Date"].astype(str).str[:7]
        g = dd.groupby("月")["Equity"].agg(["first", "last"])
        g["月次損益"] = g["last"] - g["first"]
        for j, h in enumerate(["月", "月初資産", "月末資産", "月次損益"], start=4):
            c = ws2.cell(row=31, column=j, value=h)
            c.font = F_HEAD; c.fill = PatternFill("solid", fgColor=NAVY)
        for i, (m, r) in enumerate(g.iterrows(), start=32):
            ws2.cell(row=i, column=4, value=m).font = F_BODY
            ws2.cell(row=i, column=5, value=float(r["first"])).number_format = "#,##0"
            ws2.cell(row=i, column=6, value=float(r["last"])).number_format = "#,##0"
            c = ws2.cell(row=i, column=7, value=float(r["月次損益"]))
            c.number_format = "+#,##0;-#,##0"
            c.font = Font(name="Yu Gothic", size=10, bold=True,
                          color=PROFIT if r["月次損益"] >= 0 else LOSS)

    # ============================================================ 3-5. 明細シート
    def sheet_from_df(name, df, money_cols=()):
        w = wb.create_sheet(name)
        w.append(list(df.columns))
        for c in w[1]:
            c.font = F_HEAD; c.fill = PatternFill("solid", fgColor=NAVY)
            c.alignment = CENTER
        for _, r in df.iterrows():
            w.append(list(r))
        for j, col in enumerate(df.columns, start=1):
            w.column_dimensions[get_column_letter(j)].width = \
                max(12, len(str(col)) * 2 + 2)
            if col in money_cols:
                for i in range(2, w.max_row + 1):
                    cell = w.cell(row=i, column=j)
                    try:
                        v = float(cell.value)
                        cell.number_format = "+#,##0;-#,##0"
                        cell.font = Font(name="Yu Gothic", size=10,
                                         color=PROFIT if v >= 0 else LOSS)
                    except (TypeError, ValueError):
                        pass
        w.freeze_panes = "A2"
        return w

    sheet_from_df("日次ログ", daily, money_cols=("実現損益累計", "評価損益計"))
    sheet_from_df("取引履歴", trades,
                  money_cols=("価格損益円", "スワップ円", "損益円"))
    sheet_from_df("研究ログ", research)

    # 週次ペア研究の最新結果(あれば)
    pair_scan = _read_csv(R("results", "pair_scan.csv"), [])
    if len(pair_scan):
        w = sheet_from_df("ペア研究", pair_scan)
        w.insert_rows(1)
        w.merge_cells(start_row=1, start_column=1, end_row=1,
                      end_column=len(pair_scan.columns))
        c = w.cell(row=1, column=1)
        c.value = ("週次の通貨ペア横断研究（現行2システム適用・頑健スコア順）"
                   "。土曜9時に自動更新")
        c.font = F_SECT
        w.freeze_panes = "A3"

    out = os.path.join(gd, "FXドル円_運用ダッシュボード.xlsx")
    wb.save(out)
    print(f"ダッシュボード更新: {out}")
    return out


if __name__ == "__main__":
    build()
